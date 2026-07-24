import openpyxl, json, re

SRC = "/sessions/sharp-zen-davinci/mnt/uploads/Customer Lifecycle Journey Map GS transition (version 1).xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

def col_letter(c):
    return openpyxl.utils.get_column_letter(c)

def get(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v

def band(ws, r1, r2, c):
    vals = []
    for r in range(r1, r2+1):
        v = get(ws, r, c)
        if v:
            vals.append(v)
    return vals

def fill_rgb(ws, r, c):
    cell = ws.cell(row=r, column=c)
    try:
        rgb = cell.fill.fgColor.rgb
    except Exception:
        rgb = None
    if isinstance(rgb, str) and rgb != "00000000":
        return "#" + rgb[2:]
    return None

def merge_map(ws, header_row):
    """Map column -> (title, color) for columns inside a merged range on header_row,
    using the top-left cell's title/fill for the whole span."""
    m = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= header_row <= rng.max_row:
            title = get(ws, rng.min_row, rng.min_col)
            color = fill_rgb(ws, rng.min_row, rng.min_col)
            for c in range(rng.min_col, rng.max_col+1):
                m[c] = (title, color)
    return m

def stage_ranges(ws, header_row, min_col, max_col):
    blocks = []
    cur_color = None
    cur_start = None
    cur_title = None
    mmap = merge_map(ws, header_row)
    for c in range(min_col, max_col+1):
        if c in mmap:
            title, color = mmap[c]
        else:
            color = fill_rgb(ws, header_row, c)
            title = get(ws, header_row, c)
        if color != cur_color:
            if cur_color is not None:
                blocks.append([cur_title, cur_color, cur_start, c-1])
            cur_color = color
            cur_start = c
            cur_title = title
        else:
            if title:
                cur_title = title
    if cur_color is not None:
        blocks.append([cur_title, cur_color, cur_start, max_col])
    return [b for b in blocks if b[1]]

ws1 = wb["Customer Lifecycle"]
meta1 = {
    "doc_name": get(ws1, 2, 2),
    "objective": get(ws1, 3, 2),
    "champion": get(ws1, 4, 2),
    "stakeholders": get(ws1, 5, 2),
}

stages1 = stage_ranges(ws1, 8, 2, 35)

def extract_col_sheet1(ws, c):
    event = get(ws, 11, c)
    trigger = get(ws, 13, c)
    frequency = get(ws, 14, c)
    objectives = band(ws, 16, 26, c)
    who_roles = []
    for r in range(28, 41):
        role = get(ws, r, 1)
        mark = get(ws, r, c)
        if role and mark:
            mark_s = str(mark).strip()
            note = None
            if mark_s.upper() != "X":
                note = re.sub(r"^X\s*[-\(]?\s*", "", mark_s).rstrip(")").strip()
                note = note if note else None
            who_roles.append({"role": role, "note": note})
    key_activities = band(ws, 42, 48, c)
    templates = band(ws, 50, 53, c)
    systems = []
    for r in [55, 56]:
        role = get(ws, r, 1)
        mark = get(ws, r, c)
        if role and mark:
            systems.append(role)
    celebration = get(ws, 58, c)
    metrics = band(ws, 60, 65, c)
    digital_assets = get(ws, 67, c)
    if not any([event, trigger, frequency, objectives, who_roles, key_activities, templates, systems, celebration, metrics, digital_assets]):
        return None
    return {
        "col": col_letter(c),
        "event": event,
        "trigger": trigger,
        "frequency": frequency,
        "objectives": objectives,
        "who": who_roles,
        "key_activities": key_activities,
        "templates": templates,
        "systems": systems,
        "celebration": celebration,
        "metrics": metrics,
        "digital_assets": digital_assets,
    }

sheet1_data = {"meta": meta1, "stages": []}
for title, color, c0, c1 in stages1:
    cols = []
    for c in range(c0, c1+1):
        d = extract_col_sheet1(ws1, c)
        if d:
            cols.append(d)
    if not title:
        title = "Delivery & Execution (Complex Programs)"
    sheet1_data["stages"].append({"title": title, "color": color, "columns": cols})

ws2 = wb["High Touch - Onboarding Journey"]
meta2 = {
    "doc_name": get(ws2, 2, 2),
    "objective": get(ws2, 3, 2),
    "champion": get(ws2, 4, 2),
    "stakeholders": get(ws2, 5, 2),
}
stages2 = stage_ranges(ws2, 8, 2, 11)

def extract_col_sheet2(ws, c):
    event = get(ws, 11, c)
    trigger = get(ws, 13, c)
    objectives = band(ws, 15, 23, c)
    who_roles = []
    for r in range(26, 49):
        role = get(ws, r, 1)
        mark = get(ws, r, c)
        if role and mark:
            mark_s = str(mark).strip()
            note = None
            if mark_s.upper() != "X":
                note = re.sub(r"^X\s*[-\(]?\s*", "", mark_s).rstrip(")").strip()
                note = note if note else None
            who_roles.append({"role": role, "note": note})
    key_activities = band(ws, 50, 56, c)
    templates = band(ws, 58, 61, c)
    systems = []
    for r in [63, 64, 65, 66]:
        role = get(ws, r, 1)
        mark = get(ws, r, c)
        if role and mark:
            systems.append(role)
    celebration = get(ws, 68, c)
    education = band(ws, 70, 72, c)
    metrics = band(ws, 74, 77, c)
    if not any([event, trigger, objectives, who_roles, key_activities, templates, systems, celebration, education, metrics]):
        return None
    return {
        "col": col_letter(c),
        "event": event,
        "trigger": trigger,
        "objectives": objectives,
        "who": who_roles,
        "key_activities": key_activities,
        "templates": templates,
        "systems": systems,
        "celebration": celebration,
        "education": education,
        "metrics": metrics,
    }

sheet2_data = {"meta": meta2, "stages": []}
for title, color, c0, c1 in stages2:
    cols = []
    for c in range(c0, c1+1):
        d = extract_col_sheet2(ws2, c)
        if d:
            cols.append(d)
    sheet2_data["stages"].append({"title": title, "color": color, "columns": cols})

out = {"sheet1": sheet1_data, "sheet2": sheet2_data}
with open("extracted.json", "w") as f:
    json.dump(out, f, indent=2)

print("DONE")
print("Sheet1 stages:", [(s['title'], len(s['columns'])) for s in sheet1_data['stages']])
print("Sheet2 stages:", [(s['title'], len(s['columns'])) for s in sheet2_data['stages']])
